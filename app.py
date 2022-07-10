#####Import packages#####

from email.message import Message
from timeit import repeat
import tkinter
import tkinter.messagebox
import customtkinter
from tkinter import Tk, Frame, Menu, ttk,Toplevel
from PIL import Image,ImageTk
from test import *
from tkinter.filedialog import askopenfile,asksaveasfilename,asksaveasfile
from openpyxl import load_workbook
import time
from tkinter.ttk import *
import shutil
import os
import pandas as pd
import numpy as np
import shutil
import csv


############################
#create the main window

customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"


class App(customtkinter.CTk):

    WIDTH = 780
    HEIGHT = 520

    def __init__(self):
        super().__init__()

        self.title("FIGEAC DJO-B.py")
        self.geometry(f"{App.WIDTH}x{App.HEIGHT}")
        self.protocol("WM_DELETE_WINDOW", self.on_closing)  # call .on_closing() when app gets closed

        # ============ create two frames ============

        # configure grid layout (2x1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.frame_left = customtkinter.CTkFrame(master=self,
                                                 width=180,
                                                 corner_radius=0)
        self.frame_left.grid(row=0, column=0, sticky="nswe")

        self.frame_right = customtkinter.CTkFrame(master=self)
        self.frame_right.grid(row=0, column=1, sticky="nswe", padx=20, pady=20)

        # ============ frame_left ============

        # configure grid layout (1x11)
        self.frame_left.grid_rowconfigure(0, minsize=10)   # empty row with minsize as spacing
        self.frame_left.grid_rowconfigure(5, weight=1)  # empty row as spacing
        self.frame_left.grid_rowconfigure(8, minsize=20)    # empty row with minsize as spacing
        self.frame_left.grid_rowconfigure(11, minsize=10)  # empty row with minsize as spacing

        self.label_1 = customtkinter.CTkLabel(master=self.frame_left,
                                              text="TOOLS",
                                              text_font=("Roboto Medium", -16))  # font name and size in px
        self.label_1.grid(row=1, column=0, pady=10, padx=10)

        self.button_1 = customtkinter.CTkButton(master=self.frame_left,
                                                text ='Choose DATA', 
                                                command=self.open_file)
                                               
        self.button_1.grid(row=2, column=0, pady=10, padx=20)

        self.button_2 = customtkinter.CTkButton(master=self.frame_left,
                                                text="List Of DATA",
                                                command=self.lister)
        self.button_2.grid(row=3, column=0, pady=10, padx=20)
        
        self.button_10 = customtkinter.CTkButton(master=self.frame_left,
                                                text ='Clean DATA', 
                                                command=self.read_file)
                                               
        self.button_10.grid(row=4, column=0, pady=10, padx=20)

        self.button_3 = customtkinter.CTkButton(master=self.frame_left,
                                                text="RESET",
                                                command=self.reset_app)
        self.button_3.grid(row=5, column=0, pady=10, padx=20)

        self.label_mode = customtkinter.CTkLabel(master=self.frame_left, text="Appearance Mode:")
        self.label_mode.grid(row=9, column=0, pady=0, padx=20, sticky="w")

        self.optionmenu_1 = customtkinter.CTkOptionMenu(master=self.frame_left,
                                                        values=["Light", "Dark", "System"],
                                                        command=self.change_appearance_mode)
        self.optionmenu_1.grid(row=10, column=0, pady=10, padx=20, sticky="w")

        # ============ frame_right ============

        # configure grid layout (3x7)
        self.frame_right.rowconfigure((0, 1, 2, 3), weight=1)
        self.frame_right.rowconfigure(7, weight=10)
        self.frame_right.columnconfigure((0, 1), weight=1)
        self.frame_right.columnconfigure(2, weight=0)

        self.frame_info = customtkinter.CTkFrame(master=self.frame_right)
        self.frame_info.grid(row=0, column=0, columnspan=2, rowspan=4, pady=20, padx=20, sticky="nsew")

        # ============ frame_info ============

        # configure grid layout (1x1)
        self.frame_info.rowconfigure(0, weight=1)
        self.frame_info.columnconfigure(0, weight=1)

        self.label_info_1 = customtkinter.CTkLabel(master=self.frame_info,
                                                   text="KModes clustering is,\n" +
                                                        "one of the unsupervised Machine Learning algorithms,\n" +
                                                        "that is used to cluster categorical variables" ,
                                                   height=100,
                                                   fg_color=("white", "gray38"),  # <- custom tuple-color
                                                   justify=tkinter.LEFT)
        self.label_info_1.grid(column=0, row=0, sticky="nwe", padx=5, pady=5)

        self.progressbar = customtkinter.CTkProgressBar(master=self.frame_info)
        self.progressbar.grid(row=1, column=0, sticky="ew", padx=15, pady=15)

        # ============ frame_right ============

        self.radio_var = tkinter.IntVar(value=0)

        self.label_radio_group = customtkinter.CTkLabel(master=self.frame_right,
                                                        text="CTkRadioButton Group:")
        self.label_radio_group.grid(row=0, column=2, columnspan=1, pady=20, padx=10, sticky="")

        self.radio_button_1 = customtkinter.CTkRadioButton(master=self.frame_right,
                                                           variable=self.radio_var,
                                                           value=0)
        self.radio_button_1.grid(row=1, column=2, pady=10, padx=20, sticky="n")

        self.radio_button_2 = customtkinter.CTkRadioButton(master=self.frame_right,
                                                           variable=self.radio_var,
                                                           value=1)
        self.radio_button_2.grid(row=2, column=2, pady=10, padx=20, sticky="n")

        self.radio_button_3 = customtkinter.CTkRadioButton(master=self.frame_right,
                                                           variable=self.radio_var,
                                                           value=2)
        self.radio_button_3.grid(row=3, column=2, pady=10, padx=20, sticky="n")

        self.slider_1 = customtkinter.CTkSlider(master=self.frame_right,
                                                from_=0,
                                                to=1,
                                                number_of_steps=3,
                                                command=self.progressbar.set)
        self.slider_1.grid(row=4, column=0, columnspan=2, pady=10, padx=20, sticky="we")

        self.slider_2 = customtkinter.CTkSlider(master=self.frame_right,
                                                command=self.progressbar.set)
        self.slider_2.grid(row=5, column=0, columnspan=2, pady=10, padx=20, sticky="we")

        self.switch_1 = customtkinter.CTkSwitch(master=self.frame_right,
                                                text="CTkSwitch")
        self.switch_1.grid(row=4, column=2, columnspan=1, pady=10, padx=20, sticky="we")

        self.switch_2 = customtkinter.CTkSwitch(master=self.frame_right,
                                                text="CTkSwitch")
        self.switch_2.grid(row=5, column=2, columnspan=1, pady=10, padx=20, sticky="we")
        
        self.combobox_1 = customtkinter.CTkComboBox(master=self.frame_right,
                                                    values='')
        
        self.combobox_1.grid(row=6, column=2, columnspan=1, pady=10, padx=20, sticky="we")

        self.check_box_1 = customtkinter.CTkCheckBox(master=self.frame_right,
                                                     text="CTkCheckBox")
        self.check_box_1.grid(row=6, column=0, pady=10, padx=20, sticky="w")

        self.check_box_2 = customtkinter.CTkCheckBox(master=self.frame_right,
                                                     text="CTkCheckBox")
        self.check_box_2.grid(row=6, column=1, pady=10, padx=20, sticky="w")

        self.entry = customtkinter.CTkEntry(master=self.frame_right,
                                            width=120,
                                            placeholder_text="CTkEntry")
        self.entry.grid(row=8, column=0, columnspan=2, pady=20, padx=20, sticky="we")

        self.button_5 = customtkinter.CTkButton(master=self.frame_right,
                                                text="Cluster DATA",
                                                border_width=2,  # <- custom border_width
                                                fg_color=None,  # <- no fg_color
                                                command=self.cluster_data)
        self.button_5.grid(row=8, column=2, columnspan=1, pady=20, padx=20, sticky="we")

        # set default values
        self.optionmenu_1.set("Dark")
        
        self.combobox_1.set("select file")
        self.radio_button_1.select()
        self.slider_1.set(0.2)
        self.slider_2.set(0.7)
        self.progressbar.set(0.5)
        self.switch_2.select()
        self.radio_button_3.configure(state=tkinter.DISABLED)
        self.check_box_1.configure(state=tkinter.DISABLED, text="CheckBox disabled")
        self.check_box_2.select()
        


    def button_event(self):
        print("Button pressed")

    def change_appearance_mode(self, new_appearance_mode):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def on_closing(self,event=0):
        #delete tmp files if exist
       
   
        # directory
        dir =r'C:\Users\fahmi\Desktop\clustering model\tmp'
   
        # path
        try:
            for file in os.scandir(dir):
                os.remove(file.path)
       
        except:
            pass
       
        if tkinter.messagebox.askokcancel("Quit", "Do you want to quit?"):
            self.destroy()
    
    li=[]
    
    def open_file(self):
        self.label_info_1 = customtkinter.CTkLabel(master=self.frame_info,
                                                   text="KModes clustering is,\n" +
                                                        "one of the unsupervised Machine Learning algorithms,\n" +
                                                        "that is used to cluster categorical variables" ,
                                                   height=100,
                                                   fg_color=("white", "gray38"),  # <- custom tuple-color
                                                   justify=tkinter.LEFT)
        self.label_info_1.grid(column=0, row=0, sticky="nwe", padx=5, pady=5)
        
        file = askopenfile(mode ='r', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')]) # To open the file that you want. 
    #' mode='r' ' is to tell the filedialog to read the file
    # 'filetypes=[()]' is to filter the files shown as only Excel files

        #wb = load_workbook(filename = file.name) # Load into openpyxl
        
            
        self.li.append(file.name)
            
        print(self.li)
        src_path = file.name
        dst_path = r"C:\Users\fahmi\Desktop\clustering model\tmp"
        shutil.copy(src_path, dst_path)

        pb1 = Progressbar(
        master=self.frame_right,
        orient= 'horizontal', 
        length=300, 
        mode='determinate'
        )
        pb1.grid(row=4, columnspan=3, pady=20)
        for i in range(3):
            self.update_idletasks()
            pb1['value'] += 45
            time.sleep(1)
        pb1.destroy()
        Label(self, text='File Uploaded Successfully!', foreground='green').grid(row=4, columnspan=3, pady=10)

        

        
        
        
        
              

        #wb2 = wb.active
        if file is not None:
             pass
    #Whatever you want to do with the WorkSheet
    
    #wb2.sheetname

    def save_file(self):
        file = asksaveasfile(mode ='w', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')]) # To open the file that you want.
    #' mode='w' ' is to tell the filedialog to write the file
    # 'filetypes=[()]' is to filter the files shown as only Excel files
      
        wb = load_workbook(filename = file.name) # Load into openpyxl
       
        wb4 = wb.active
        if file is not None:
            pass
    def uploadFiles(self):
        pb1 = Progressbar(
        master=self.frame_right,
        orient= 'horizontal', 
        length=300, 
        mode='determinate'
        )
        pb1.grid(row=4, columnspan=3, pady=20)
        for i in range(5):
            self.update_idletasks()
            pb1['value'] += 20
            time.sleep(1)
        pb1.destroy()
        Label(self, text='File Uploaded Successfully!', foreground='green').grid(row=4, columnspan=3, pady=10)
        
        
    ### fUNCTION TO TREE THE DATA 
    newli=[]
    
    def lister(self):
        new_window=Toplevel(self.master)
        new_window.geometry("600x400+50+50")
        new_window.attributes('-alpha', 0.85)
        new_window.title("List of Files")
        new_window.resizable(False,False)
        new_window.configure(background='#f5f5f5')
        new_window.focus_force()
        new_window.grab_set()
        new_window.transient(self.master)

        tree =ttk.Treeview(new_window, columns=("ID","File Name", "File Size", "File Type"), height=200, show='headings',  
                            style='mystyle.Treeview', selectmode='browse')

        tree_scroll = ttk.Scrollbar(new_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=tree_scroll.set)
        tree_scroll.pack(side="right", fill="y")
        tree.pack(side="left", fill="both", expand=True)
        
        tree['columns'] = ('ID','name', 'size', 'type')
        tree.column("ID", width=50)
        tree.column("name", width=100)
        tree.column("size", width=100)
        tree.column("type", width=100)
        tree.heading("ID", text="ID")
        tree.heading("name", text="File Name")
        tree.heading("size", text="File Size")
        tree.heading("type", text="File Type")
        count=0
        for i in self.li:
            j=i.split("/")
            h=j[5].split(".")
        
            
            tree.insert("", "end", text='',values=(count,h[0], os.path.getsize(i), os.path.splitext(i)[1]))
            tree.pack()
            count+=1
        self.newli.append(h[0])
        self.newli = list(dict.fromkeys(self.newli))
        print(self.newli)         
        self.combobox_1 = customtkinter.CTkComboBox(master=self.frame_right,
                                                    values=self.newli,)
        
        self.combobox_1.grid(row=6, column=2, columnspan=1, pady=10, padx=20, sticky="we")
        

############ FUNCTION TO CLEAN THE DATA ###############

    def read_file(self):
        from itertools import groupby
        c=self.combobox_1.get()
        path=r"C:\Users\fahmi\Desktop\clustering model\tmp" +"\\"+ c + ".xlsx"
        path1=r"C:\Users\fahmi\Desktop\clustering model\tmp"+"\\"+"df.csv"
        read_file = pd.read_excel(path)
        read_file.to_csv (path1, index = None,header=True)
        df = pd.read_csv(path1)
        
        df.groupby(['RefsPieces'])
        df= df.drop(df.columns[1],axis=1)
        df= df.drop(df.columns[1],axis=1)
        df= df.drop(df.columns[2],axis=1)
        df= df.drop(df.columns[2],axis=1)
        df= df.drop(df.columns[2],axis=1)
        df= df.drop(df.columns[2],axis=1)

        

        dict={}
        from collections import OrderedDict, defaultdict
        studentdict=df.to_dict('split')
        #studentdict=studentdict.values()
        s=studentdict['data']

        list_of_lists = s
        d={}
        all_values = [list[0] for list in list_of_lists]
        unique_values = set(all_values)

        group_list = []
        for value in unique_values:
            this_group = []
            for list in list_of_lists:
                if list[0] == value:
                    this_group.append(list[1])
                    d[list[0]]= this_group
            group_list.append(this_group)
  


        dff=pd.DataFrame(d.items())
        clusterpath=r"C:\Users\fahmi\Desktop\clustering model\tmp"+"\\"+"datareg.csv"
        
        dff.to_csv(clusterpath, index = None,header=True)
        datareg = pd.read_csv(clusterpath)
        datareg = datareg.set_index('0')

        datareg = datareg.replace('\[', '', regex=True)
        datareg = datareg.replace('\]', '', regex=True)
        datareg=datareg['1'].str.split(r",", expand=True)
        datareg.to_csv(clusterpath, index = True ,header=True)
        Label(self, text='DATA Cleaned up Successfully!', foreground='green').grid(row=4, columnspan=3, pady=10)
        
#########################################################################################################

    def cluster_data(self):
        import joblib as joblib
        import pandas as pd 
        import numpy as np
        clusterpath=r"C:\Users\fahmi\Desktop\clustering model\tmp"+"\\"+"datareg.csv"
        cls=joblib.load('modeloo.sav')
        dataclean=pd.read_csv(clusterpath)
        dataclean = dataclean.replace(np.nan,"", regex=True)
        #predict dataclean with cls
        clusters=cls.predict(dataclean)

        dataclean.insert(0, "Cluster", clusters, True)

        c0=dataclean[dataclean['Cluster']==0]
        c1=dataclean[dataclean['Cluster']==1]
        c2=dataclean[dataclean['Cluster']==2]
        c3=dataclean[dataclean['Cluster']==3]
        
       
        filename0 = asksaveasfile(initialdir='/', title='Save File', filetypes=(('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv'),('Text Files', 'txt.*'), ('All Files', '*.*')))
        filename1 = asksaveasfile(initialdir='/', title='Save File', filetypes=(('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv'),('Text Files', 'txt.*'), ('All Files', '*.*')))
        filename2 = asksaveasfile(initialdir='/', title='Save File', filetypes=(('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv'),('Text Files', 'txt.*'), ('All Files', '*.*')))
        filename3 = asksaveasfile(initialdir='/', title='Save File', filetypes=(('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv'),('Text Files', 'txt.*'), ('All Files', '*.*')))
        c0.to_excel(filename0.name, index = True ,header=True)
        c1.to_excel(filename1.name, index = True ,header=True)
        c2.to_excel(filename2.name, index = True ,header=True)
        c3.to_excel(filename3.name, index = True ,header=True)


        tkinter.messagebox.showinfo("Success", "Data has been clustered")
        Message='File 1 saved as : ' + filename0.name + '\n' + 'File 2 saved as : ' + filename1.name + '\n' + 'File 3 saved as : ' + filename2.name + '\n' + 'File 4 saved as : ' + filename3.name
        
        self.label_info_1 = customtkinter.CTkLabel(master=self.frame_info,
                                                   text= Message,
                                                   height=100,
                                                   fg_color=("white", "gray38"),  # <- custom tuple-color
                                                   justify=tkinter.LEFT)
        self.label_info_1.grid(column=0, row=0, sticky="nwe", padx=5, pady=5)
        
        
    #create function that reset the app 
    def reset_app(self):
        #delete tmp files if exist
        # directory
        dir =r'C:\Users\fahmi\Desktop\clustering model\tmp' 
        # path
        try:
            for file in os.scandir(dir):
                os.remove(file.path)  
                self.button_3.config(state="disabled") 
        except:
            pass
    
          

#create toolbar Menu  and add it to the main window
class Example(Frame):
    
        def __init__(self):
            super().__init__()

            self.initUI()


        def initUI(self):

            self.master.title("FIGEAC DJO-B")

            menubar = Menu(self.master)
            self.master.config(menu=menubar)

            fileMenu = Menu(menubar)
            fileMenu.add_command(label="Exit", command=self.onExit)
            menubar.add_cascade(label="File", menu=fileMenu)
            
            editMenu = Menu(menubar)
            editMenu.add_command(label="Undo")
            editMenu.add_separator()
            editMenu.add_command(label="Cut")
            editMenu.add_command(label="Copy")
            editMenu.add_command(label="Paste")
            editMenu.add_separator()
            editMenu.add_command(label="Select All")
            menubar.add_cascade(label="Edit", menu=editMenu)
            
            helpMenu = Menu(menubar)
            helpMenu.add_command(label="About", command=self.onAbout)
            menubar.add_cascade(label="Help", menu=helpMenu)
            
            
            
           
            


        def onExit(self):

            self.quit()   

        
        ## FUNCTION TO SHOW ABOUT DIALOG     
        def onAbout(self):
            new_window=Toplevel(self.master)
            new_window.geometry("800x400+50+50")
            new_window.attributes('-alpha', 0.7)
            new_window.title("About")
            new_window.resizable(False,False)
            lbl=customtkinter.CTkLabel(master=new_window,
                                                   text=
                                                   '''Auteur : FAHMI DJOBBI \n \n \t'''+
'''Les outils coupants représentent une dépense importante pour toutes les entreprises d'usinage.\n'''+
'Chaque référence est fabriquée en utilisant des outils de coupe spécifiques par une machine spécialisée.\n'+

'En fait, pour fabriquer une pièce quelconque on doit ajouter Les outils qui le nécessitent dans le pont de machine.\n' +
'Ensuite, nous ajoutons la pièce à sa place et démarrons la machine pour commencer la création.\n'+

'''En parallèle, les outils de coupe ont été référencés et classés grâce au logiciel GEDIX, qui a pour but de gérer nos outils de coupe,\n'''+ 
'''de standardiser et de rationaliser notre parc d'outils tout en capitalisant sur notre savoir-faire et qui a permis de savoir 
sur quel produit les outils étaient appelés.\n'''+

'''Le problème qui se pose est qu''à chaque fois qu'une nouvelle référence est ajoutée, les outils de coupe doivent être remplacés par
d''autres et comme l'ajout est effectué par la main-d''œuvre, cela peut prendre beaucoup de temps.\n'''+

'''D'autre part, l'entreprise peut subir une baisse de productivité, surtout si toutes les pièces nécessitent des outils différents.
Dans ce cas, en plus de la baisse de productivité, l'entreprise perd du temps et de l'argent.\n'''+

'''Cette problématique nous a motivé à proposer une solution qui permettrait de regrouper les pièces qui utilisent 
les mêmes outils de coupeet prédire pour les nouvelles pièces à quel groupe elles appartiennent. 
D''où le nom de note projet « Optimisation de la gestion des outils coupants »\n '''+

'''Cette solution est basée sur une base de données qui contient les informations 
sur les outils de coupe et les pièces qui les utilisent.\n''' ,


                                                   
                                                   height=150,
                                                   fg_color=("black"),  #<- custom tuple-color
                                                   justify="left",
                                                   )

            lbl.grid(row=0, column=0, columnspan=2, pady=10, padx=20, sticky="we")
            lbl.configure(font=("Arial", 9, "bold"))

             
 
 
 
 
 
 
 ######################### END  APP ###################################               

if __name__ == "__main__":
    app = App()
    app1=Example()
    app.mainloop()
    
    
    
#fahmi Djobbi